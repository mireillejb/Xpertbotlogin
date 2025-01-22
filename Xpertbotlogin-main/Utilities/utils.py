import inspect
import logging
import os
import json
from openpyxl import Workbook, load_workbook
import datetime


class AutomationLogger:
    @staticmethod
    def automation(log_level=logging.DEBUG):  
        logger_name = inspect.stack()[1][3]
        logger = logging.getLogger(logger_name)
        logger.setLevel(log_level)

        today = datetime.date.today()
        formatted_date = today.strftime("%Y-%m-%d")  # Format the date as needed

        log_file_name = f"Logs/automation_{formatted_date}.log"
        os.makedirs(os.path.dirname(log_file_name), exist_ok=True)

        file_handler = logging.FileHandler(log_file_name, mode='a')
        formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
        file_handler.setFormatter(formatter)

        logger.addHandler(file_handler)
        return logger


    def get_newest_excel_file(folder_path, sheet=None):
        files = [f for f in os.listdir(folder_path) if f.endswith(".xlsx")]
        print("Files in the folder:", files)

        if not files:
            raise FileNotFoundError(f"No Excel files found in the specified folder: {folder_path}")

        # Sort files by creation time (descending order)
        sorted_files = sorted(files, key=lambda f: os.path.getctime(os.path.join(folder_path, f)), reverse=True)
        print("Sorted files by creation time:", sorted_files)

        latest_file = sorted_files[0]
        print("Latest file:", latest_file)

        file_path = os.path.join(folder_path, latest_file)
        print("Selected file path:", file_path)

        try:
            wb = load_workbook(filename=file_path)
        except Exception as e:
            raise FileNotFoundError(f"Error opening the file: {e}")

        datalist = []

        if sheet is None:
            sheet = wb.sheetnames[0]  # Use the first sheet by default if sheet is not specified

        sh = wb[sheet]
        row_ct = sh.max_row
        col_ct = sh.max_column

        for i in range(2, row_ct + 1):
            row = {}
            for j in range(1, col_ct + 1):
                column_name = sh.cell(row=1, column=j).value
                column_value = sh.cell(row=i, column=j).value
                row[column_name] = column_value
            datalist.append(row)

        return datalist